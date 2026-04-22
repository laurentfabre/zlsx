"""Matching openpyxl writer benchmark — 1000 rows × 10 cols + header style."""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def main() -> None:
    out = sys.argv[1]
    wb = Workbook(write_only=True)
    sheet = wb.create_sheet("Bench")

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(fgColor="FF1E3A8A", patternType="solid")
    header_align = Alignment(horizontal="center")

    # Header row with cell-level styles.
    from openpyxl.cell import WriteOnlyCell
    header_names = ["Name", "Amount", "Share", "Qty", "Active", "Code",
                    "A", "B", "C", "D"]
    header_cells = []
    for name in header_names:
        c = WriteOnlyCell(sheet, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        header_cells.append(c)
    sheet.append(header_cells)

    # Data rows — no per-cell styles (openpyxl write-only makes
    # styling each body cell very expensive; mirror zlsx where body
    # cells reference a shared numFmt style but we don't per-cell-
    # attach full Font/Fill/Alignment).
    for i in range(1000):
        sheet.append([
            f"row_{i}",
            100.0 + i * 1.5,
            i / 1000.0,
            i,
            i % 2 == 0,
            f"CODE_{i:x}",
            i * 7,
            i * 0.1,
            "x",
            None,
        ])

    wb.save(out)


if __name__ == "__main__":
    main()
