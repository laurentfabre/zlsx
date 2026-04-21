"""Same benchmark workload for openpyxl (read-only + values-only — fastest openpyxl mode)."""
import sys
from openpyxl import load_workbook

def main() -> None:
    path = sys.argv[1]
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    n_rows = n_str = n_int = n_num = n_bool = n_empty = 0
    for row in ws.iter_rows(values_only=True):
        n_rows += 1
        for c in row:
            if c is None:
                n_empty += 1
            elif isinstance(c, bool):
                n_bool += 1
            elif isinstance(c, int):
                n_int += 1
            elif isinstance(c, float):
                n_num += 1
            else:
                n_str += 1
    wb.close()
    print(f"rows={n_rows} str={n_str} int={n_int} num={n_num} bool={n_bool} empty={n_empty}")

if __name__ == "__main__":
    main()
